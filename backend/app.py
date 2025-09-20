from flask import Flask, render_template, request, redirect, url_for, flash
from datetime import datetime
import os
import io
import json
import openpyxl
from openpyxl import Workbook, load_workbook
import shutil
import tempfile
from google.oauth2 import service_account
from googleapiclient.discovery import build
from googleapiclient.http import MediaIoBaseDownload, MediaIoBaseUpload

app = Flask(__name__)
app.secret_key = "supersecretkey"

# -------------------------
# Dosya yolları (geçici dizin)
# -------------------------
EXCEL_FILE_LOCAL = os.path.join(tempfile.gettempdir(), "lojistik.xlsx")
EXCEL_FILE_ONEDRIVE = os.path.join(tempfile.gettempdir(), "OneDrive_lojistik.xlsx")

# -------------------------
# Google Drive ayarları
# -------------------------
EXCEL_FILE_DRIVE_ID = "1Rvg3nQkHsVjh9QicnU5ViYvzJm1EwO8T"
SCOPES = ['https://www.googleapis.com/auth/drive.file']

# -------------------------
# Google Drive servis fonksiyonları
# -------------------------
def get_drive_service():
    credentials_json = os.environ.get("GOOGLE_CREDENTIALS_JSON")
    if not credentials_json:
        raise Exception("GOOGLE_CREDENTIALS_JSON environment variable bulunamadı")
    
    credentials_info = json.loads(credentials_json)
    if "private_key" in credentials_info and "\\n" in credentials_info["private_key"]:
        credentials_info["private_key"] = credentials_info["private_key"].replace("\\n", "\n")
    
    creds = service_account.Credentials.from_service_account_info(credentials_info, scopes=SCOPES)
    service = build('drive', 'v3', credentials=creds)
    return service

def download_excel(service, file_id):
    request = service.files().get_media(fileId=file_id)
    fh = io.BytesIO()
    downloader = MediaIoBaseDownload(fh, request)
    done = False
    while not done:
        _, done = downloader.next_chunk()
    fh.seek(0)
    try:
        wb = openpyxl.load_workbook(fh)
    except Exception as e:
        raise Exception(f"Google Drive Excel yüklenemedi: {e}")
    return wb

def upload_excel(service, file_id, wb):
    fh = io.BytesIO()
    wb.save(fh)
    fh.seek(0)
    media = MediaIoBaseUpload(fh, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    service.files().update(fileId=file_id, media_body=media).execute()

# -------------------------
# Ana route
# -------------------------
@app.route("/", methods=["GET", "POST"])
def form():
    if request.method == "POST":
        try:
            # Form verilerini al, boşsa varsayılan ata
            tarih = request.form.get("tarih") or datetime.now().strftime("%Y-%m-%d")
            iscikissaat = request.form.get("iscikissaat") or "00:00"
            plaka = request.form.get("plaka") or "Bilinmiyor"
            cikiskm = float(request.form.get("cikiskm") or 0)
            kumgirissaat = request.form.get("kumgirissaat") or "00:00"
            giriskm = float(request.form.get("giriskm") or 0)
            kumcikissaat = request.form.get("kumcikissaat") or "00:00"
            isletmegiriskm = float(request.form.get("isletmegiriskm") or 0)
            isletmegirissaat = request.form.get("isletmegirissaat") or "00:00"
            farkkm = giriskm - cikiskm
            uretici = request.form.get("uretici") or "Bilinmiyor"
            ureticikm = float(request.form.get("ureticikm") or 0)
            tonaj = int(request.form.get("tonaj") or 0)

            # -------------------------
            # Lokal Excel kaydı
            # -------------------------
            try:
                if not os.path.exists(EXCEL_FILE_LOCAL):
                    wb = Workbook()
                    ws = wb.active
                    ws.append([
                        "tarih","iscikissaat","plaka","cikiskm","kumgirissaat",
                        "giriskm","kumcikissaat","isletmegiriskm","isletmegirissaat",
                        "farkkm","uretici","ureticikm","tonaj"
                    ])
                    wb.save(EXCEL_FILE_LOCAL)

                wb = load_workbook(EXCEL_FILE_LOCAL)
                ws = wb.active
                ws.append([tarih, iscikissaat, plaka, cikiskm, kumgirissaat,
                           giriskm, kumcikissaat, isletmegiriskm, isletmegirissaat,
                           farkkm, uretici, ureticikm, tonaj])
                wb.save(EXCEL_FILE_LOCAL)

                # OneDrive kopyası
                shutil.copy(EXCEL_FILE_LOCAL, EXCEL_FILE_ONEDRIVE)
            except Exception as e:
                flash(f"Excel kaydetme hatası: {e}", "danger")

            # -------------------------
            # Google Drive kaydı
            # -------------------------
            try:
                service = get_drive_service()
                wb_drive = download_excel(service, EXCEL_FILE_DRIVE_ID)
                ws_drive = wb_drive.active
                ws_drive.append([tarih, iscikissaat, plaka, cikiskm, kumgirissaat,
                                 giriskm, kumcikissaat, isletmegiriskm, isletmegirissaat,
                                 farkkm, uretici, ureticikm, tonaj])
                upload_excel(service, EXCEL_FILE_DRIVE_ID, wb_drive)
            except Exception as e:
                flash(f"Google Drive’a kaydetme hatası: {e}", "warning")

            flash("Kayıt başarıyla eklendi!", "success")
            return redirect(url_for("form"))

        except Exception as e:
            flash(f"Hata oluştu: {e}", "danger")
            return redirect(url_for("form"))

    return render_template("form.html")

# -------------------------
# Uygulama başlatma
# -------------------------
if __name__ == "__main__":
    port = int(os.environ.get("PORT", 5000))
    app.run(host="0.0.0.0", port=port, debug=True)
