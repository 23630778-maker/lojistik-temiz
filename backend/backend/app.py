from flask import Flask, render_template, request, redirect, url_for, flash
from datetime import datetime
import os
from openpyxl import Workbook, load_workbook
import shutil
import json
from google.oauth2 import service_account
from googleapiclient.discovery import build
from googleapiclient.http import MediaIoBaseUpload, MediaIoBaseDownload
import io
import pandas as pd  # ⬅️ verileri tablo olarak göstermek için eklendi

app = Flask(__name__)
app.secret_key = "supersecretkey"

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
EXCEL_FILE_LOCAL = os.path.join(BASE_DIR, "lojistik.xlsx")
EXCEL_FILE_ONEDRIVE = os.path.join(BASE_DIR, "OneDrive_lojistik.xlsx")

# Google Drive ayarları
EXCEL_FILE_DRIVE_ID = "1Rvg3nQkHsVjh9QicnU5ViYvzJm1EwO8T"  # Drive'daki dosya ID
JSON_PATH = os.path.join(BASE_DIR, "credentials.json")  # credentials.json dosya yolu
SCOPES = ["https://www.googleapis.com/auth/drive.file"]

def get_drive_service():
    with open(JSON_PATH, "r", encoding="utf-8") as f:
        creds_info = json.load(f)

    if "private_key" in creds_info:
        creds_info["private_key"] = creds_info["private_key"].replace("\\n", "\n")

    creds = service_account.Credentials.from_service_account_info(creds_info, scopes=SCOPES)
    return build("drive", "v3", credentials=creds)

def download_excel(service, file_id):
    try:
        request = service.files().get_media(fileId=file_id)
        fh = io.BytesIO()
        downloader = MediaIoBaseDownload(fh, request)
        done = False
        while not done:
            _, done = downloader.next_chunk()
        fh.seek(0)
        return load_workbook(fh)
    except Exception:
        wb = Workbook()
        ws = wb.active
        ws.append([
            "tarih","iscikissaat","plaka","cikiskm","kumgirissaat",
            "giriskm","kumcikissaat","isletmegiriskm","isletmegirissaat",
            "farkkm","uretici","ureticikm","tonaj"
        ])
        return wb

def upload_excel(service, file_id, wb):
    try:
        fh = io.BytesIO()
        wb.save(fh)
        fh.seek(0)
        media = MediaIoBaseUpload(
            fh,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            resumable=False
        )
        service.files().update(fileId=file_id, media_body=media).execute()
    except Exception as e:
        print(f"[Google Drive Upload Hatası] {e}")

@app.route("/", methods=["GET", "POST"])
def form():
    if request.method == "POST":
        try:
            tarih = request.form.get("tarih") or datetime.now().strftime("%Y-%m-%d")
            iscikissaat = request.form.get("iscikissaat") or "00:00"
            plaka = request.form.get("plaka") or "Bilinmiyor"
            cikiskm = float(request.form.get("cikiskm") or 0)
            kumgirissaat = request.form.get("kumgirissaat") or "00:00"
            giriskm = float(request.form.get("giriskm") or 0)
            kumcikissaat = request.form.get("kumcikissaat") or "00:00"
            isletmegiriskm = float(request.form.get("isletmegiriskm") or 0)
            isletmegirissaat = request.form.get("isletmegirissaat") or "00:00"
            farkkm = giriskm - cikiskm
            uretici = request.form.get("uretici") or "Bilinmiyor"
            ureticikm = float(request.form.get("ureticikm") or 0)
            tonaj = int(request.form.get("tonaj") or 0)

            # Lokal Excel kaydı
            if not os.path.exists(EXCEL_FILE_LOCAL):
                wb = Workbook()
                ws = wb.active
                ws.append([
                    "tarih","iscikissaat","plaka","cikiskm","kumgirissaat",
                    "giriskm","kumcikissaat","isletmegiriskm","isletmegirissaat",
                    "farkkm","uretici","ureticikm","tonaj"
                ])
                wb.save(EXCEL_FILE_LOCAL)

            wb = load_workbook(EXCEL_FILE_LOCAL)
            ws = wb.active
            ws.append([
                tarih, iscikissaat, plaka, cikiskm, kumgirissaat,
                giriskm, kumcikissaat, isletmegiriskm, isletmegirissaat,
                farkkm, uretici, ureticikm, tonaj
            ])
            wb.save(EXCEL_FILE_LOCAL)

            # OneDrive kopyası
            shutil.copy(EXCEL_FILE_LOCAL, EXCEL_FILE_ONEDRIVE)

            # Google Drive kaydı
            try:
                service = get_drive_service()
                wb_drive = download_excel(service, EXCEL_FILE_DRIVE_ID)
                ws_drive = wb_drive.active
                ws_drive.append([
                    tarih, iscikissaat, plaka, cikiskm, kumgirissaat,
                    giriskm, kumcikissaat, isletmegiriskm, isletmegirissaat,
                    farkkm, uretici, ureticikm, tonaj
                ])
                upload_excel(service, EXCEL_FILE_DRIVE_ID, wb_drive)
            except Exception as e:
                print(f"[Google Drive Genel Hatası] {e}")

            flash("Kayıt başarıyla eklendi!", "success")
            return redirect(url_for("veriler"))  # ⬅️ Kayıt sonrası veriler sayfasına yönlendir
        except Exception as e:
            flash(f"Hata oluştu: {e}", "danger")
            return redirect(url_for("form"))

    return render_template("form.html")

@app.route("/veriler")
def veriler():
    if os.path.exists(EXCEL_FILE_LOCAL):
        df = pd.read_excel(EXCEL_FILE_LOCAL)
        tablo_html = df.to_html(classes="table table-striped", index=False)
    else:
        tablo_html = "<p>Henüz veri yok.</p>"

    return render_template("veriler.html", tablo=tablo_html)

if __name__ == "__main__":
    port = int(os.environ.get("PORT", 5000))
    app.run(host="0.0.0.0", port=port, debug=True)
